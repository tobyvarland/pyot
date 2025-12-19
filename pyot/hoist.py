import logging
import os
import shlex
import shutil
import socket
import subprocess
import csv
from abc import ABC, abstractmethod
from datetime import date, datetime
from pathlib import Path
from zoneinfo import ZoneInfo
from typing import Optional, List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import CellIsRule


class HoistAggregator:
    """Aggregates hoist CSV data into a single master CSV file.

    Reads multiple hoist data CSV files as defined in the provided
    HoistAggregationConfig, normalizes and filters rows, derives additional
    fields (such as station type and duration), sorts records chronologically,
    and writes the consolidated result to a single output CSV file.

    Attributes:
        config (HoistAggregationConfig): Configuration object defining aggregation
            behavior, source files, station type mappings, and output location.
    """

    OUTPUT_DATE_FORMAT = "%Y-%m-%d %H:%M:%S"

    def __init__(self, config) -> None:
        """Initialize the hoist aggregator.

        Args:
            config (HoistAggregationConfig): Aggregation configuration.
        """
        self.config = config

    def run(self) -> None:
        """Execute the hoist aggregation process.

        If aggregation is disabled in the configuration, this method returns
        immediately. Otherwise, it collects rows from all configured source
        files, sorts them by load timestamp, and writes the consolidated output.
        """
        if not self.config.enabled:
            return

        rows = self._collect_rows()
        rows.sort(key=lambda r: r["_sort"])
        self._write_output(rows)

    def _collect_rows(self) -> list[dict]:
        """Collect normalized rows from all configured hoist data files.

        Iterates over each HoistAggregationSpec in the configuration, reads the
        corresponding CSV file, and processes each row.

        Returns:
            list[dict]: List of normalized row dictionaries.
        """
        rows: list[dict] = []

        for spec in self.config.files:
            with spec.path.open(newline="", encoding="utf-8") as f:
                reader = csv.reader(f)
                next(reader, None)  # skip header row

                for raw in reader:
                    row = self._process_row(raw, spec)
                    if row is not None:
                        rows.append(row)

        return rows

    def _process_row(
        self,
        raw: list[str],
        spec,
    ) -> Optional[dict]:
        """Process a single CSV row for a given hoist specification.

        Applies filtering rules, parses timestamps, derives duration and station
        type, and returns a normalized row dictionary. Rows that do not meet
        validation or filtering criteria are discarded.

        Args:
            raw (list[str]): Raw CSV row values.
            spec (HoistAggregationSpec): Specification describing how to interpret
                the row.

        Returns:
            Optional[dict]: Normalized row dictionary, or None if the row should
                be skipped.
        """
        shop_order = self._safe_get(raw, spec.indices.get("shop_order"))
        if shop_order in ("", "0", "111"):
            return None
        
        if "date_in" in spec.indices and "time_in" in spec.indices:
            loaded_dt = self._parse_timestamp(
                self._safe_get(raw, spec.indices.get("date_in")),
                self._safe_get(raw, spec.indices.get("time_in")),
            )
        else:
            try:
                date_part, time_part = self._safe_get(raw, spec.indices.get("dt_in")).split()
                month, day, year = map(int, date_part.split("/"))
                reformatted_date = f"{year % 100:02d}{month:02d}{day:02d}"
                reformatted_time = time_part.replace(":", "")
                loaded_dt = self._parse_timestamp(
                    reformatted_date,
                    reformatted_time,
                )
            except Exception:
                loaded_dt = None
        
        if "date_out" in spec.indices and "time_out" in spec.indices:
            unloaded_dt = self._parse_timestamp(
                self._safe_get(raw, spec.indices.get("date_out")),
                self._safe_get(raw, spec.indices.get("time_out")),
            )
        else:
            try:
                date_part, time_part = self._safe_get(raw, spec.indices.get("dt_out")).split()
                month, day, year = map(int, date_part.split("/"))
                reformatted_date = f"{year % 100:02d}{month:02d}{day:02d}"
                reformatted_time = time_part.replace(":", "")
                unloaded_dt = self._parse_timestamp(
                    reformatted_date,
                    reformatted_time,
                )
            except Exception:
                unloaded_dt = None

        if not loaded_dt or not unloaded_dt:
            return None

        station_number_str = self._safe_get(raw, spec.indices.get("station"))
        station_number = self._to_int(station_number_str)

        station_type = ""
        if station_number is not None:
            station_type = self.config.station_types.get(
                (spec.lane, station_number),
                ""
            )

        if not station_type:
            return None
        
        actual_ah = ""
        if station_type == "PLATE" and "actual_ah" not in spec.indices.values():
            target_ah = self._safe_get(raw, spec.indices.get("target_ah"))
            ah_pct = self._safe_get(raw, spec.indices.get("ah_pct"))
            if target_ah and ah_pct:
                try:
                    actual_ah_value = float(target_ah) * float(ah_pct)
                    actual_ah = f"{actual_ah_value:.3f}"
                except ValueError:
                    actual_ah = ""
        else:
            actual_ah = self._safe_get(raw, spec.indices.get("actual_ah"))

        return {
            "Hoist #": spec.hoist,
            "Lane Number": spec.lane,
            "Station Number": station_number_str,
            "Station Type": station_type,

            "Date/Time Loaded": self._format_datetime(loaded_dt),
            "Date/Time Unloaded": self._format_datetime(unloaded_dt),
            "Duration": self._format_duration(loaded_dt, unloaded_dt),

            "Customer": self._safe_get(raw, spec.indices.get("customer")),
            "Part ID": self._safe_get(raw, spec.indices.get("part")),
            "Shop Order": shop_order,
            "Load Number": self._safe_get(raw, spec.indices.get("load")),
            "Barrel Number": self._safe_get(raw, spec.indices.get("barrel")),

            "Target Amp Hours":
                self._safe_get(raw, spec.indices.get("target_ah"))
                if station_type == "PLATE" else "",

            "Actual Amp Hours":
                actual_ah
                if station_type == "PLATE" else "",

            "Amp Hours Percent":
                self._safe_get(raw, spec.indices.get("ah_pct"))
                if station_type == "PLATE" else "",

            "Barrel Speed": self._safe_get(raw, spec.indices.get("barrel_speed")),
            "Target Weight": self._safe_get(raw, spec.indices.get("target_weight")),
            "Actual Weight": self._safe_get(raw, spec.indices.get("actual_weight")),

            "_sort": unloaded_dt,
        }

    def _write_output(self, rows: list[dict]) -> None:
        """Write aggregated hoist data to the configured output CSV file.

        Args:
            rows (list[dict]): Normalized and sorted row dictionaries.
        """
        with self.config.output_file.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=[
                    "Hoist #",
                    "Lane Number",
                    "Station Number",
                    "Station Type",
                    "Date/Time Loaded",
                    "Date/Time Unloaded",
                    "Duration",
                    "Customer",
                    "Part ID",
                    "Shop Order",
                    "Load Number",
                    "Barrel Number",
                    "Target Amp Hours",
                    "Actual Amp Hours",
                    "Amp Hours Percent",
                    "Barrel Speed",
                    "Target Weight",
                    "Actual Weight",
                ],
            )
            writer.writeheader()

            for row in rows:
                row.pop("_sort", None)
                writer.writerow(row)

    @staticmethod
    def _safe_get(row: list[str], index: Optional[int]) -> str:
        """Safely retrieve and strip a value from a CSV row.

        Args:
            row (list[str]): CSV row values.
            index (Optional[int]): Column index.

        Returns:
            str: Stripped value or empty string if unavailable.
        """
        if index is None or index < 0 or index >= len(row):
            return ""
        return row[index].strip()

    @staticmethod
    def _to_int(value: str) -> Optional[int]:
        """Convert a string to an integer if possible.

        Args:
            value (str): String value.

        Returns:
            Optional[int]: Parsed integer or None if conversion fails.
        """
        try:
            return int(value)
        except ValueError:
            return None

    @staticmethod
    def _parse_timestamp(date_str: str, time_str: str) -> Optional[datetime]:
        """Parse a YYMMDD / HHMMSS timestamp into a datetime.

        Args:
            date_str (str): Date string in YYMMDD format.
            time_str (str): Time string in HHMMSS format.

        Returns:
            Optional[datetime]: Parsed datetime or None if parsing fails.
        """
        if not date_str or not time_str:
            return None
        try:
            while len(time_str) < 6:
                time_str = "0" + time_str
            return datetime.strptime(f"{date_str}{time_str}", "%y%m%d%H%M%S")
        except ValueError:
            return None

    def _format_datetime(self, dt: datetime) -> str:
        """Format a datetime for CSV output.

        Args:
            dt (datetime): Datetime to format.

        Returns:
            str: Formatted datetime string.
        """
        return dt.strftime(self.OUTPUT_DATE_FORMAT)

    @staticmethod
    def _format_duration(start: datetime, end: datetime) -> str:
        """Format the duration between two datetimes as H:MM:SS.

        Args:
            start (datetime): Start time.
            end (datetime): End time.

        Returns:
            str: Duration string.
        """
        seconds = int((end - start).total_seconds())
        sign = "-" if seconds < 0 else ""
        seconds = abs(seconds)
        h, r = divmod(seconds, 3600)
        m, s = divmod(r, 60)
        return f"{sign}{h}:{m:02d}:{s:02d}"
    

class HoistExcelExporter:
    """Exports aggregated hoist data from CSV to a formatted Excel workbook.

    Reads a consolidated hoist CSV file and produces an Excel (.xlsx) file with
    predefined formatting, filters, column widths, and data types suitable for
    analysis and reporting.
    
    Attributes:
        csv_path (Path): Path to the source CSV file.
        xlsx_path (Path): Path to the output Excel file.
    """

    """Columns containing datetime values."""
    DATETIME_COLUMNS = {"Date/Time Loaded", "Date/Time Unloaded"}

    """Columns containing floating-point numeric values."""
    FLOAT_COLUMNS = {
        "Target Amp Hours",
        "Actual Amp Hours",
        "Target Weight",
        "Actual Weight",
    }

    """Columns containing integer values."""
    INTEGER_COLUMNS = {"Hoist #", "Lane Number", "Station Number", "Shop Order", "Load Number", "Barrel Number", "Barrel Speed"}

    """Columns that should be treated as text."""
    TEXT_COLUMNS = {"Customer", "Part ID", "Station Type"}

    """Columns that should be treated as percentages."""
    PERCENTAGE_COLUMNS = {"Amp Hours Percent"}

    """Excel table style to apply to the worksheet."""
    TABLE_STYLE = "TableStyleMedium1"

    def __init__(self, csv_path: Path, xlsx_path: Path) -> None:
        """Initialize the Excel exporter.

        Args:
            csv_path (Path): Path to the consolidated CSV file.
            xlsx_path (Path): Path to the Excel file to be written.
        """
        self.csv_path = csv_path
        self.xlsx_path = xlsx_path

    def write(self) -> None:
        """Generate the Excel workbook from the CSV source.

        Reads the CSV file, applies formatting and presentation rules, and writes
        the resulting Excel workbook to disk.
        """
        rows = self._read_csv()
        workbook = self._create_workbook(rows)
        workbook.save(self.xlsx_path)

    def _read_csv(self) -> list[list[str]]:
        """Read all rows from the CSV source file.

        Returns:
            list[list[str]]: List of CSV rows.
        """
        with self.csv_path.open(newline="", encoding="utf-8") as f:
            return list(csv.reader(f))

    def _create_workbook(self, rows: list[list[str]]):
        """Create and populate an Excel workbook from CSV rows.

        Args:
            rows (list[list[str]]): CSV data rows.

        Returns:
            Workbook: Populated Excel workbook.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Hoist Data"

        header = rows[0]
        ws.append(header)

        for row in rows[1:]:
            converted_row = []
            for col_idx, value in enumerate(row):
                col_name = header[col_idx] if col_idx < len(header) else ""
                converted_row.append(self._convert_value(value, col_name))
            ws.append(converted_row)

        self._apply_filters(ws)
        self._apply_table(ws)
        self._format_rows(ws)
        self._format_columns(ws, header)
        self._apply_conditional_formatting(ws, header)

        return wb
    
    def _convert_value(self, value: str, col_name: str):
        """Convert a CSV string value to the appropriate Python type.

        Args:
            value (str): The string value from CSV.
            col_name (str): The column name to determine type.

        Returns:
            The converted value (int, float, datetime, or str).
        """
        if not value or value.strip() == "":
            return ""

        try:
            if col_name in self.DATETIME_COLUMNS:
                return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
            elif col_name in self.INTEGER_COLUMNS:
                return int(value)
            elif col_name in self.FLOAT_COLUMNS or col_name in self.PERCENTAGE_COLUMNS:
                return float(value)
            elif col_name == "Duration":
                is_negative = value.startswith("-")
                clean_value = value.lstrip("-")
                parts = clean_value.split(":")
                hours = int(parts[0])
                minutes = int(parts[1])
                seconds = int(parts[2])
                total_seconds = hours * 3600 + minutes * 60 + seconds
                duration_days = total_seconds / 86400
                return -duration_days if is_negative else duration_days
        except (ValueError, TypeError):
            pass

        return value

    def _apply_filters(self, ws) -> None:
        """Enable column filters and freeze the header row.

        Args:
            ws: Worksheet to modify.
        """
        ws.freeze_panes = "A2"

    def _apply_table(self, ws) -> None:
        """Apply Excel table styling to the worksheet.

        Args:
            ws: Worksheet to modify.
        """
        table = Table(displayName="HoistAggregation", ref=ws.dimensions)
        style = TableStyleInfo(
            name=self.TABLE_STYLE,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    def _format_rows(self, ws) -> None:
        """Apply row height to all rows.

        Args:
            ws: Worksheet to modify.
        """
        for row_num in range(1, ws.max_row + 1):
            ws.row_dimensions[row_num].height = 21

    def _format_columns(self, ws, header: list[str]) -> None:
        """Apply column widths and number formatting.

        Args:
            ws: Worksheet to modify.
            header (list[str]): Column header names.
        """
        # Define column-specific settings
        column_config = {
            "Hoist #": {
                "width": 10,
                "alignment": "center",
                "number_format": "0"
            },
            "Lane Number": {
                "width": 15,
                "alignment": "center",
                "number_format": "0"
            },
            "Station Number": {
                "width": 15,
                "alignment": "center",
                "number_format": "0"
            },
            "Station Type": {
                "width": 15,
                "alignment": "center",
                "number_format": "@"
            },
            "Date/Time Loaded": {
                "width": 20,
                "alignment": "left",
                "number_format": "mm/dd/yyyy hh:mm:ss"
            },
            "Date/Time Unloaded": {
                "width": 20,
                "alignment": "left",
                "number_format": "mm/dd/yyyy hh:mm:ss"
            },
            "Duration": {
                "width": 12,
                "alignment": "center",
                "number_format": "[h]:mm:ss"
            },
            "Customer": {
                "width": 12,
                "alignment": "left",
                "number_format": "@"
            },
            "Part ID": {
                "width": 25,
                "alignment": "left",
                "number_format": "@"
            },
            "Shop Order": {
                "width": 12,
                "alignment": "center",
                "number_format": "0"
            },
            "Load Number": {
                "width": 12,
                "alignment": "center",
                "number_format": "0"
            },
            "Barrel Number": {
                "width": 14,
                "alignment": "center",
                "number_format": "0"
            },
            "Target Amp Hours": {
                "width": 18,
                "alignment": "right",
                "number_format": "0.0"
            },
            "Actual Amp Hours": {
                "width": 18,
                "alignment": "right",
                "number_format": "0.0"
            },
            "Amp Hours Percent": {
                "width": 18,
                "alignment": "right",
                "number_format": "0.00%"
            },
            "Barrel Speed": {
                "width": 12,
                "alignment": "center",
                "number_format": "0"
            },
            "Target Weight": {
                "width": 14,
                "alignment": "right",
                "number_format": "0.0"
            },
            "Actual Weight": {
                "width": 14,
                "alignment": "right",
                "number_format": "0.0"
            },
        }

        for col_idx, col_name in enumerate(header, start=1):
            column_letter = get_column_letter(col_idx)
            
            # Get configuration for this column
            config = column_config.get(col_name, {
                "width": 12,
                "alignment": "left",
                "number_format": "General"
            })

            # Set column width
            ws.column_dimensions[column_letter].width = config["width"]

            # Apply formatting to cells
            ws[column_letter][0].alignment = Alignment(
                horizontal=config["alignment"],
                vertical='center'
            )
            cells = ws[column_letter][1:]
            for cell in cells:
                cell.number_format = config["number_format"]
                cell.alignment = Alignment(
                    horizontal=config["alignment"],
                    vertical='center'
                )

    def _apply_conditional_formatting(self, ws, header: list[str]) -> None:
        """Apply conditional formatting rules.

        Args:
            ws: Worksheet to modify.
            header (list[str]): Column header names.
        """
        # Find the Amp Hours Percent column
        try:
            amp_hours_col_idx = header.index("Amp Hours Percent") + 1
            amp_hours_col_letter = get_column_letter(amp_hours_col_idx)
            
            # Define the range (skip header row, go to max row)
            range_str = f"{amp_hours_col_letter}2:{amp_hours_col_letter}{ws.max_row}"
            
            # Red bold font for values < 90%
            red_bold_font = Font(color="FF0000", bold=True)
            rule_low = CellIsRule(
                operator='lessThan',
                formula=['0.9'],
                font=red_bold_font
            )
            ws.conditional_formatting.add(range_str, rule_low)
            
            # Red bold font for values > 110%
            rule_high = CellIsRule(
                operator='greaterThan',
                formula=['1.1'],
                font=red_bold_font
            )
            ws.conditional_formatting.add(range_str, rule_high)
            
        except ValueError:
            # Column not found, skip conditional formatting
            pass